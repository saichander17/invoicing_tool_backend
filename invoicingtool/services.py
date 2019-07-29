from invoicingtool.models import Invoice, UploadedInvoiceFile
from invoicingtool.serializers import InvoiceSerializer, UploadedInvoiceFileSerializer
from django.utils import timezone
from background_task import background
import openpyxl, csv
from django.core.files.storage import FileSystemStorage
from django.db import transaction

class InvoicesService(object):
  """docstring for ClassName"""
  # def __init__(self, arg):
  #   super(ClassName, self).__init__()
  #   self.arg = arg
    
  def fetch(self, offset=0, limit=10):
    if offset:
      offset = int(offset)
    if limit:
      limit = int(limit)
    if limit==None and offset==None:
      return Invoice.objects.all()
    elif limit==None:
      return Invoice.objects.all()[offset:limit]
    else:
      return Invoice.objects.all()[offset:offset+limit]


class InvoiceFinderService(object):
  
  def find(self, id):
    return Invoice.objects.get(id=id)
    

class InvoiceCreatorService(object):
  def create(self, data):
    serializer = InvoiceSerializer(data={
      'customer_info': data['customer_info'],
      'product_info': data['product_info'],
      'pricing': data['pricing'],
      'invoice_path': None,
      'created_at': timezone.now(),
      'updated_at': timezone.now()
      })
    if serializer.is_valid():
      serializer.save()
      return serializer.data
    else:
      return False

class UploadedInvoiceFileFinderService(object):
  def find(self, id):
    return UploadedInvoiceFile.objects.get(id=id)
    
    
    

class InvoiceUploaderService(object):
  def upload(self, file):
    # This method should save the file from the args into appropriate location
    # We can have third party services as well which gets called from here. 
    # eg. You can upload the files to S3 from here and delete the local versions. You could have S3UploaderService for this purpose
    # For now, I'm using localstorage
    destination_folder = "invoicingtool/uploaded_files/" + "---".join(timezone.now().strftime('%B %d, %Y, %I:%M %p').split(" "))
    file_path = destination_folder + file.name
    destination = open(file_path, 'wb+')
    for chunk in file.chunks():
      destination.write(chunk)
      destination.close()
    # Invoke the background worker with the uploaded file id to create inoices
    serializer = UploadedInvoiceFileSerializer(data={
      'path': file_path,
      'percentage_processed': 0
      })
    if serializer.is_valid():
      serializer.save()
      BulkInvoiceCreatorService().create(serializer.data['id'])
      return {'success': True, 'file_id': serializer.data['id']}
    else:
      return {'success': False}


class BulkInvoiceCreatorService(object):
  # ToDo: Add transaction
  # @background(schedule=10)
  def create(self, file_id):
    file_headers = ["s.no", "customername", "productname", "quantity", "amount", "taxpercent", "discountpercent"]
    uploaded_file = UploadedInvoiceFile.objects.get(id=file_id)
    if uploaded_file==None:
      return False
    file_path = uploaded_file.path
    if(file_path.endswith('csv')):
      self.parseCSV(uploaded_file)
    else:
      self.parseXl(uploaded_file)
    return True

  def parseCSV(self, uploaded_file):
    file_headers = ["s.no", "customername", "productname", "quantity", "amount", "taxpercent", "discountpercent"]
    file_path = uploaded_file.path
    f = open(file_path)
    uploaded_file.status='processing'
    uploaded_file.save()
    # reader = csv.reader(f)
    reader = csv.DictReader(f,fieldnames=file_headers)
    next(reader, None)
    next(reader, None)
    next(reader, None)
    headers = next(reader, None)
    old_obj = None
    sub_total = 0
    creator_service = InvoiceCreatorService()
    access_obj = {}
    for idx,val in enumerate(file_headers):
      access_obj[val] = val
    for row in reader:
      old_obj, sub_total = self.create_old_obj(old_obj, row, access_obj, sub_total, creator_service)
    if old_obj!=None:
      self.calculateTotalValues(old_obj, sub_total)
      if not creator_service.create(old_obj):
        uploaded_file.status='failure'
        uploaded_file.save()
        raise Exception('Some error occurred while saving an invoice')
    uploaded_file.status='success'
    uploaded_file.percentage_processed = 100
    uploaded_file.save()
  
  def parseXl(self, uploaded_file):
    file_headers = ["s.no", "customername", "productname", "quantity", "amount", "taxpercent", "discountpercent"]
    file_path = uploaded_file.path
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames
    worksheet = wb[sheets[0]]
    old_obj = None
    sub_total = 0
    creator_service = InvoiceCreatorService()
    access_obj = {}
    for idx,val in enumerate(file_headers):
      access_obj[val] = idx
    for row in worksheet.iter_rows(min_row=2, values_only=True):
      old_obj, sub_total = self.create_old_obj(old_obj, row, access_obj, sub_total, creator_service)
    if old_obj!=None:
      self.calculateTotalValues(old_obj, sub_total)
      if not creator_service.create(old_obj):
        uploaded_file.status='failure'
        uploaded_file.save()
        raise Exception('Some error occurred while saving an invoice')
    uploaded_file.status='success'
    uploaded_file.percentage_processed = 100
    uploaded_file.save()

  def create_old_obj(self, old_obj, row, access_obj, sub_total, creator_service):
    file_headers = ["s.no", "customername", "productname", "quantity", "amount", "taxpercent", "discountpercent"]
    if(row[access_obj["s.no"]]!=None and row[access_obj["s.no"]]!=""):
      try:
        float(row[access_obj["s.no"]])
      except ValueError:
        return old_obj,sub_total
        # print(row[access_obj["s.no"]])
        # raise Exception('Invalid invoice uploaded')
    if(row[access_obj["s.no"]]!=None and row[access_obj["s.no"]]!=""):
      if old_obj!=None:
        self.calculateTotalValues(old_obj, sub_total)
        if not creator_service.create(old_obj):
          uploaded_file.status='failure'
          uploaded_file.save()
          raise Exception('Some error occurred while saving an invoice')
      sub_total = 0
      old_obj = {"customer_info": {}, "product_info": {}, "pricing": {}}
    # breakpoint()
    old_obj["customer_info"]["name"] = old_obj["customer_info"].get("name") or row[access_obj['customername']]
    old_obj["product_info"] = old_obj["product_info"] or []
    old_obj["product_info"].append({"name": row[access_obj["productname"]], "amount": float(row[access_obj['amount']]), "quantity": float(row[access_obj['quantity']])})
    old_obj["pricing"] = old_obj["pricing"] or {"tax_percent": float(row[access_obj['taxpercent']]), "discount_percent": float(row[access_obj['discountpercent']])}
    sub_total += float(row[access_obj['amount']])
    return old_obj, sub_total

  def calculateTotalValues(self, obj, sub_total):
    obj["pricing"]["subtotal"] = sub_total
    obj["pricing"]["tax"] = (sub_total*obj["pricing"]["tax_percent"])/100.0
    obj["pricing"]["discount"] = (sub_total*obj["pricing"]["discount_percent"])/100.0
    obj["pricing"]["total_amount"] = sub_total - obj["pricing"]["discount"] + obj["pricing"]["tax"]

